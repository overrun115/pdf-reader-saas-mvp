#!/usr/bin/env python3
"""
Enhanced Excel Analyzer with Real Value
Análisis útil de Excel con extracción de datos y insights accionables
"""

import openpyxl
import pandas as pd
import json
from typing import Dict, List, Any, Tuple, Optional
import numpy as np
from datetime import datetime
import re
import logging

logger = logging.getLogger(__name__)

class EnhancedExcelAnalyzer:
    """Enhanced Excel analyzer that extracts actionable insights and real data"""
    
    def __init__(self):
        self.numeric_threshold = 0.7  # 70% numeric to consider a column numeric
        self.date_patterns = [
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',
            r'\d{1,2}-\w{3}-\d{2,4}'
        ]
    
    def analyze_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Comprehensive Excel analysis with real data extraction"""
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Basic file info
            sheet_names = workbook.sheetnames
            total_sheets = len(sheet_names)
            
            # Analyze each sheet
            sheets_analysis = []
            extracted_data = {}
            insights = []
            
            for sheet_name in sheet_names:
                sheet_analysis = self._analyze_sheet(workbook[sheet_name], sheet_name)
                sheets_analysis.append(sheet_analysis)
                
                # Extract actual data from important sheets
                if sheet_analysis['importance_score'] > 0.7:
                    extracted_data[sheet_name] = self._extract_sheet_data(workbook[sheet_name])
            
            workbook.close()
            
            # Generate intelligent insights
            insights = self._generate_insights(sheets_analysis, extracted_data)
            
            # Detect file purpose and type
            file_classification = self._classify_excel_file(sheets_analysis, extracted_data)
            
            # Calculate practical value score
            value_score = self._calculate_value_score(sheets_analysis, extracted_data)
            
            return {
                "file_classification": file_classification,
                "total_sheets": total_sheets,
                "value_score": value_score,
                "sheets_analysis": sheets_analysis,
                "extracted_data": extracted_data,
                "insights": insights,
                "actionable_recommendations": self._get_actionable_recommendations(
                    file_classification, extracted_data, insights
                ),
                "export_options": self._get_export_options(extracted_data),
                "data_summary": self._create_data_summary(extracted_data)
            }
            
        except Exception as e:
            logger.error(f"Error in enhanced Excel analysis: {str(e)}")
            raise
    
    def _analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual sheet with real insights"""
        
        # Get data range
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 1 or max_col == 1:
            return {
                "name": sheet_name,
                "status": "empty_or_minimal",
                "importance_score": 0,
                "data_type": "empty"
            }
        
        # Sample data for analysis (first 100 rows, 20 columns max)
        sample_rows = min(max_row, 100)
        sample_cols = min(max_col, 20)
        
        # Extract headers (first row)
        headers = []
        for col in range(1, sample_cols + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")
        
        # Analyze column types and data patterns
        column_analysis = self._analyze_columns(sheet, sample_rows, sample_cols, headers)
        
        # Detect data patterns and structure
        data_patterns = self._detect_data_patterns(sheet, sample_rows, sample_cols, column_analysis)
        
        # Calculate importance score
        importance_score = self._calculate_sheet_importance(
            max_row, max_col, column_analysis, data_patterns, sheet_name
        )
        
        # Get sample data (first 5 rows)
        sample_data = self._get_sample_data(sheet, min(6, sample_rows), sample_cols, headers)
        
        return {
            "name": sheet_name,
            "dimensions": {"rows": max_row, "columns": max_col},
            "headers": headers,
            "column_analysis": column_analysis,
            "data_patterns": data_patterns,
            "importance_score": importance_score,
            "sample_data": sample_data,
            "data_density": self._calculate_data_density(sheet, sample_rows, sample_cols),
            "estimated_records": max_row - 1 if max_row > 1 else 0
        }
    
    def _analyze_columns(self, sheet, max_row: int, max_col: int, headers: List[str]) -> List[Dict[str, Any]]:
        """Analyze each column's data type and characteristics"""
        column_analysis = []
        
        for col in range(1, max_col + 1):
            analysis = {
                "name": headers[col-1] if col <= len(headers) else f"Column_{col}",
                "index": col,
                "data_type": "unknown",
                "numeric_count": 0,
                "text_count": 0,
                "date_count": 0,
                "empty_count": 0,
                "unique_values": set(),
                "sample_values": [],
                "is_calculated": False,
                "statistics": {}
            }
            
            numeric_values = []
            
            # Analyze data in column (skip header row)
            for row in range(2, min(max_row + 1, 50)):  # Sample first 48 rows
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                
                if value is None:
                    analysis["empty_count"] += 1
                elif isinstance(value, (int, float)):
                    analysis["numeric_count"] += 1
                    numeric_values.append(float(value))
                    analysis["unique_values"].add(value)
                elif isinstance(value, str):
                    analysis["text_count"] += 1
                    analysis["unique_values"].add(value)
                    
                    # Check if it looks like a date
                    if self._is_date_like(value):
                        analysis["date_count"] += 1
                else:
                    analysis["date_count"] += 1  # datetime objects
                
                # Keep sample values
                if len(analysis["sample_values"]) < 5 and value is not None:
                    analysis["sample_values"].append(str(value)[:50])  # Truncate long values
                
                # Check for formulas (basic detection)
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    analysis["is_calculated"] = True
            
            # Determine primary data type
            total_non_empty = analysis["numeric_count"] + analysis["text_count"] + analysis["date_count"]
            
            if total_non_empty == 0:
                analysis["data_type"] = "empty"
            elif analysis["numeric_count"] / max(total_non_empty, 1) > self.numeric_threshold:
                analysis["data_type"] = "numeric"
                # Calculate statistics for numeric columns
                if numeric_values:
                    analysis["statistics"] = {
                        "min": min(numeric_values),
                        "max": max(numeric_values),
                        "avg": sum(numeric_values) / len(numeric_values),
                        "sum": sum(numeric_values),
                        "count": len(numeric_values)
                    }
            elif analysis["date_count"] / max(total_non_empty, 1) > 0.5:
                analysis["data_type"] = "date"
            else:
                analysis["data_type"] = "text"
            
            analysis["unique_count"] = len(analysis["unique_values"])
            del analysis["unique_values"]  # Remove to save space
            
            column_analysis.append(analysis)
        
        return column_analysis
    
    def _detect_data_patterns(self, sheet, max_row: int, max_col: int, column_analysis: List[Dict]) -> Dict[str, Any]:
        """Detect meaningful data patterns"""
        
        # Financial patterns
        has_financial_headers = any(
            any(keyword in col["name"].lower() for keyword in 
                ["price", "cost", "revenue", "profit", "amount", "total", "sum", "value", "money", "dollar", "€", "$"])
            for col in column_analysis
        )
        
        # Date/time patterns
        has_temporal_data = any(
            col["data_type"] == "date" or 
            any(keyword in col["name"].lower() for keyword in ["date", "time", "month", "year", "day"])
            for col in column_analysis
        )
        
        # Sales/business patterns
        has_business_data = any(
            any(keyword in col["name"].lower() for keyword in 
                ["customer", "product", "order", "sales", "quantity", "client", "item", "sku"])
            for col in column_analysis
        )
        
        # Inventory patterns
        has_inventory_data = any(
            any(keyword in col["name"].lower() for keyword in 
                ["stock", "inventory", "quantity", "units", "warehouse", "supplier"])
            for col in column_analysis
        )
        
        # Look for calculated totals
        has_calculations = any(col["is_calculated"] for col in column_analysis)
        
        # Detect table structure
        numeric_cols = sum(1 for col in column_analysis if col["data_type"] == "numeric")
        text_cols = sum(1 for col in column_analysis if col["data_type"] == "text")
        
        return {
            "has_financial_data": has_financial_headers,
            "has_temporal_data": has_temporal_data,
            "has_business_data": has_business_data,
            "has_inventory_data": has_inventory_data,
            "has_calculations": has_calculations,
            "numeric_columns": numeric_cols,
            "text_columns": text_cols,
            "data_structure": self._determine_data_structure(column_analysis),
            "likely_purpose": self._determine_sheet_purpose(column_analysis)
        }
    
    def _extract_sheet_data(self, sheet) -> Dict[str, Any]:
        """Extract actual usable data from important sheets"""
        
        # Get dimensions
        max_row = min(sheet.max_row, 1000)  # Limit to first 1000 rows
        max_col = min(sheet.max_column, 50)  # Limit to first 50 columns
        
        if max_row <= 1 or max_col <= 1:
            return {"status": "no_data"}
        
        # Extract headers
        headers = []
        for col in range(1, max_col + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(str(header) if header is not None else f"Column_{col}")
        
        # Extract data rows
        data_rows = []
        for row in range(2, min(max_row + 1, 102)):  # First 100 data rows
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                
                # Convert to appropriate type
                if cell_value is None:
                    row_data.append(None)
                elif isinstance(cell_value, (int, float)):
                    row_data.append(cell_value)
                else:
                    row_data.append(str(cell_value))
            
            # Only include rows that have some data
            if any(val is not None and val != "" for val in row_data):
                data_rows.append(row_data)
        
        # Create summary statistics for numeric columns
        numeric_summaries = {}
        for col_idx, header in enumerate(headers):
            values = []
            for row in data_rows:
                if col_idx < len(row) and isinstance(row[col_idx], (int, float)):
                    values.append(row[col_idx])
            
            if len(values) > 0:
                numeric_summaries[header] = {
                    "count": len(values),
                    "sum": sum(values),
                    "avg": sum(values) / len(values),
                    "min": min(values),
                    "max": max(values)
                }
        
        return {
            "status": "extracted",
            "headers": headers,
            "data_rows": data_rows[:20],  # First 20 rows for preview
            "total_data_rows": len(data_rows),
            "numeric_summaries": numeric_summaries,
            "extraction_timestamp": datetime.now().isoformat()
        }
    
    def _generate_insights(self, sheets_analysis: List[Dict], extracted_data: Dict) -> List[str]:
        """Generate actionable insights from the analysis"""
        insights = []
        
        # Analyze across all sheets
        important_sheets = [s for s in sheets_analysis if s.get('importance_score', 0) > 0.7]
        
        if important_sheets:
            insights.append(f"Found {len(important_sheets)} sheet(s) with significant data content")
            
            # Data type insights
            total_numeric_cols = sum(s.get('data_patterns', {}).get('numeric_columns', 0) for s in important_sheets)
            if total_numeric_cols > 5:
                insights.append(f"Rich numerical data detected ({total_numeric_cols} numeric columns) - suitable for analysis")
            
            # Business data insights
            business_sheets = [s for s in important_sheets if s.get('data_patterns', {}).get('has_business_data')]
            if business_sheets:
                insights.append("Business/sales data detected - can extract customer and product information")
            
            # Financial insights
            financial_sheets = [s for s in important_sheets if s.get('data_patterns', {}).get('has_financial_data')]
            if financial_sheets:
                insights.append("Financial data detected - can calculate totals, trends, and summaries")
                
                # Look for actual financial calculations
                for sheet_name, data in extracted_data.items():
                    if data.get('numeric_summaries'):
                        for col_name, stats in data['numeric_summaries'].items():
                            if 'total' in col_name.lower() or 'sum' in col_name.lower():
                                insights.append(f"Total value in '{col_name}': {stats['sum']:,.2f}")
            
            # Temporal insights
            temporal_sheets = [s for s in important_sheets if s.get('data_patterns', {}).get('has_temporal_data')]
            if temporal_sheets:
                insights.append("Date/time data found - can analyze trends over time")
            
            # Data volume insights
            total_records = sum(s.get('estimated_records', 0) for s in important_sheets)
            if total_records > 100:
                insights.append(f"Large dataset detected ({total_records:,} records) - suitable for data analysis")
        else:
            insights.append("Limited useful data found - file may be template or have minimal content")
        
        return insights
    
    def _get_actionable_recommendations(self, classification: str, extracted_data: Dict, insights: List[str]) -> List[Dict[str, str]]:
        """Generate specific actionable recommendations"""
        recommendations = []
        
        # Based on file classification
        if "financial" in classification.lower():
            recommendations.append({
                "action": "Generate Financial Report",
                "description": "Create automated summary of revenue, expenses, and profit margins",
                "value": "Save hours of manual calculation"
            })
            
        if "sales" in classification.lower() or "customer" in classification.lower():
            recommendations.append({
                "action": "Customer Analysis",
                "description": "Extract customer data and sales patterns for CRM integration",
                "value": "Improve customer relationship management"
            })
            
        if "inventory" in classification.lower():
            recommendations.append({
                "action": "Stock Level Analysis",
                "description": "Monitor inventory levels and identify reorder points",
                "value": "Optimize inventory management"
            })
        
        # Based on extracted data
        for sheet_name, data in extracted_data.items():
            if data.get('status') == 'extracted' and data.get('numeric_summaries'):
                recommendations.append({
                    "action": f"Export {sheet_name} Data",
                    "description": f"Convert {data['total_data_rows']} rows to CSV/JSON for analysis",
                    "value": "Enable data integration with other systems"
                })
        
        # Always useful recommendations
        if extracted_data:
            recommendations.append({
                "action": "Data Validation",
                "description": "Check for inconsistencies, duplicates, and data quality issues",
                "value": "Ensure data accuracy for decision making"
            })
            
            recommendations.append({
                "action": "Automated Reporting",
                "description": "Set up regular reports based on this data structure",
                "value": "Save time on recurring analysis tasks"
            })
        
        return recommendations
    
    def _get_export_options(self, extracted_data: Dict) -> List[Dict[str, str]]:
        """Get available export options based on extracted data"""
        options = []
        
        for sheet_name, data in extracted_data.items():
            if data.get('status') == 'extracted':
                options.append({
                    "format": "CSV",
                    "sheet": sheet_name,
                    "description": f"Export {sheet_name} as CSV file ({data['total_data_rows']} rows)",
                    "use_case": "Import into databases, analysis tools, or other spreadsheet applications"
                })
                
                options.append({
                    "format": "JSON",
                    "sheet": sheet_name,
                    "description": f"Export {sheet_name} as JSON file for API integration",
                    "use_case": "Integrate with web applications, databases, or data processing pipelines"
                })
                
                if data.get('numeric_summaries'):
                    options.append({
                        "format": "Summary Report",
                        "sheet": sheet_name,
                        "description": f"Generate executive summary of {sheet_name} data",
                        "use_case": "Quick insights for management and stakeholders"
                    })
        
        return options
    
    def _create_data_summary(self, extracted_data: Dict) -> Dict[str, Any]:
        """Create executive summary of all extracted data"""
        if not extracted_data:
            return {"status": "no_data"}
        
        total_records = sum(data.get('total_data_rows', 0) for data in extracted_data.values() 
                           if data.get('status') == 'extracted')
        
        # Collect all numeric summaries
        all_numeric_data = {}
        for sheet_name, data in extracted_data.items():
            if data.get('numeric_summaries'):
                for col_name, stats in data['numeric_summaries'].items():
                    key = f"{sheet_name}.{col_name}"
                    all_numeric_data[key] = stats
        
        # Find the most significant numbers
        significant_totals = []
        for key, stats in all_numeric_data.items():
            if stats['sum'] > 1000:  # Significant amounts
                significant_totals.append({
                    "description": key,
                    "total": stats['sum'],
                    "average": stats['avg'],
                    "count": stats['count']
                })
        
        # Sort by total value
        significant_totals.sort(key=lambda x: x['total'], reverse=True)
        
        return {
            "status": "summary_available",
            "total_records": total_records,
            "sheets_with_data": len([d for d in extracted_data.values() if d.get('status') == 'extracted']),
            "significant_totals": significant_totals[:5],  # Top 5 most significant
            "data_types_found": list(set(
                col_type for data in extracted_data.values() 
                if data.get('headers') 
                for col_type in ['numeric', 'text', 'date']  # Simplified for summary
            ))
        }
    
    # Helper methods
    def _is_date_like(self, value: str) -> bool:
        """Check if a string looks like a date"""
        for pattern in self.date_patterns:
            if re.search(pattern, value):
                return True
        return False
    
    def _calculate_sheet_importance(self, max_row: int, max_col: int, 
                                  column_analysis: List[Dict], data_patterns: Dict, 
                                  sheet_name: str) -> float:
        """Calculate how important/useful this sheet is"""
        score = 0
        
        # Size factor (more data = more important)
        data_volume = max_row * max_col
        if data_volume > 100:
            score += 0.3
        if data_volume > 1000:
            score += 0.2
        
        # Data diversity (mix of data types is good)
        unique_types = len(set(col["data_type"] for col in column_analysis))
        score += unique_types * 0.1
        
        # Business relevance
        if data_patterns.get('has_business_data'):
            score += 0.3
        if data_patterns.get('has_financial_data'):
            score += 0.3
        if data_patterns.get('has_calculations'):
            score += 0.2
        
        # Sheet name relevance
        important_names = ['data', 'sales', 'customers', 'revenue', 'inventory', 'main', 'summary']
        if any(name in sheet_name.lower() for name in important_names):
            score += 0.2
        
        return min(score, 1.0)  # Cap at 1.0
    
    def _get_sample_data(self, sheet, max_row: int, max_col: int, headers: List[str]) -> List[Dict[str, Any]]:
        """Get sample data for preview"""
        sample = []
        
        for row in range(2, min(max_row + 1, 7)):  # First 5 data rows
            row_data = {}
            for col in range(1, min(max_col + 1, 6)):  # First 5 columns
                header = headers[col-1] if col <= len(headers) else f"Col_{col}"
                value = sheet.cell(row=row, column=col).value
                
                if value is not None:
                    if isinstance(value, (int, float)):
                        row_data[header] = value
                    else:
                        # Truncate long text
                        row_data[header] = str(value)[:50] + "..." if len(str(value)) > 50 else str(value)
                else:
                    row_data[header] = None
            
            if any(v is not None for v in row_data.values()):
                sample.append(row_data)
        
        return sample
    
    def _calculate_data_density(self, sheet, max_row: int, max_col: int) -> float:
        """Calculate how much of the sheet contains actual data"""
        total_cells = max_row * max_col
        if total_cells == 0:
            return 0
        
        filled_cells = 0
        for row in range(1, min(max_row + 1, 50)):  # Sample first 50 rows
            for col in range(1, min(max_col + 1, 20)):  # Sample first 20 cols
                if sheet.cell(row=row, column=col).value is not None:
                    filled_cells += 1
        
        sample_size = min(50, max_row) * min(20, max_col)
        return filled_cells / sample_size if sample_size > 0 else 0
    
    def _determine_data_structure(self, column_analysis: List[Dict]) -> str:
        """Determine the overall structure of the data"""
        if not column_analysis:
            return "empty"
        
        numeric_ratio = sum(1 for col in column_analysis if col["data_type"] == "numeric") / len(column_analysis)
        text_ratio = sum(1 for col in column_analysis if col["data_type"] == "text") / len(column_analysis)
        
        if numeric_ratio > 0.7:
            return "primarily_numeric"
        elif text_ratio > 0.7:
            return "primarily_text"
        elif numeric_ratio > 0.3 and text_ratio > 0.3:
            return "mixed_data"
        else:
            return "structured_table"
    
    def _determine_sheet_purpose(self, column_analysis: List[Dict]) -> str:
        """Determine what this sheet is likely used for"""
        headers = [col["name"].lower() for col in column_analysis]
        
        # Financial indicators
        financial_keywords = ["revenue", "profit", "cost", "price", "amount", "total", "budget"]
        if any(keyword in " ".join(headers) for keyword in financial_keywords):
            return "financial_data"
        
        # Sales indicators
        sales_keywords = ["customer", "product", "order", "sales", "quantity", "client"]
        if any(keyword in " ".join(headers) for keyword in sales_keywords):
            return "sales_data"
        
        # Inventory indicators
        inventory_keywords = ["stock", "inventory", "warehouse", "supplier", "sku"]
        if any(keyword in " ".join(headers) for keyword in inventory_keywords):
            return "inventory_data"
        
        # HR indicators
        hr_keywords = ["employee", "salary", "department", "position", "hire"]
        if any(keyword in " ".join(headers) for keyword in hr_keywords):
            return "hr_data"
        
        return "general_data"
    
    def _classify_excel_file(self, sheets_analysis: List[Dict], extracted_data: Dict) -> str:
        """Classify the overall purpose of the Excel file"""
        purposes = [sheet.get('data_patterns', {}).get('likely_purpose', 'general') 
                   for sheet in sheets_analysis]
        
        # Count purpose types
        purpose_counts = {}
        for purpose in purposes:
            purpose_counts[purpose] = purpose_counts.get(purpose, 0) + 1
        
        # Get dominant purpose
        if purpose_counts:
            dominant_purpose = max(purpose_counts.items(), key=lambda x: x[1])[0]
            
            # Enhance classification based on extracted data
            if extracted_data and any(data.get('numeric_summaries') for data in extracted_data.values()):
                if dominant_purpose == "financial_data":
                    return "financial_analysis_workbook"
                elif dominant_purpose == "sales_data":
                    return "sales_performance_tracker"
                elif dominant_purpose == "inventory_data":
                    return "inventory_management_system"
                else:
                    return f"{dominant_purpose}_workbook"
            else:
                return f"{dominant_purpose}_template"
        
        return "general_spreadsheet"
    
    def _calculate_value_score(self, sheets_analysis: List[Dict], extracted_data: Dict) -> float:
        """Calculate overall value/usefulness score of the file"""
        if not sheets_analysis:
            return 0
        
        # Base score from sheet importance
        avg_importance = sum(sheet.get('importance_score', 0) for sheet in sheets_analysis) / len(sheets_analysis)
        score = avg_importance * 0.4
        
        # Bonus for extracted data
        if extracted_data:
            sheets_with_data = sum(1 for data in extracted_data.values() if data.get('status') == 'extracted')
            score += (sheets_with_data / len(sheets_analysis)) * 0.3
        
        # Bonus for numeric analysis capability
        numeric_summaries_count = sum(1 for data in extracted_data.values() 
                                     if data.get('numeric_summaries'))
        if numeric_summaries_count > 0:
            score += 0.2
        
        # Bonus for business relevance
        business_sheets = sum(1 for sheet in sheets_analysis 
                             if sheet.get('data_patterns', {}).get('has_business_data'))
        if business_sheets > 0:
            score += 0.1
        
        return min(score, 1.0)