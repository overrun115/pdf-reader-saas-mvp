#!/usr/bin/env python3
"""
Excel Intelligence Service - Análisis inteligente de hojas de cálculo Excel
Parte de la Fase 2 de la expansión de inteligencia documental
"""

import logging
import asyncio
import re
import io
from typing import Dict, List, Any, Optional, Tuple, Union
from dataclasses import dataclass
from datetime import datetime, date
import statistics
import numpy as np
import pandas as pd

# Excel processing libraries
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import xlrd
import xlwings as xw

# Data analysis and validation
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA

logger = logging.getLogger(__name__)

@dataclass
class ExcelWorksheet:
    """Información de hoja de Excel"""
    name: str
    index: int
    dimensions: Tuple[int, int]  # (rows, cols)
    data_range: str
    has_headers: bool
    data_types: Dict[str, str]
    statistics: Dict[str, Any]

@dataclass
class ExcelFormula:
    """Fórmula de Excel detectada"""
    cell: str
    formula: str
    formula_type: str
    dependencies: List[str]
    is_array_formula: bool

@dataclass
class DataModel:
    """Modelo de datos detectado en Excel"""
    model_type: str  # financial, inventory, schedule, report, etc.
    key_metrics: List[str]
    relationships: List[Dict[str, Any]]
    complexity_score: float
    confidence: float

class ExcelIntelligenceService:
    """
    Servicio de inteligencia para archivos Excel
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Patrones para detectar tipos de datos
        self.data_patterns = {
            'currency': r'^\$?[\d,]+\.?\d*$',
            'percentage': r'^\d+\.?\d*%$',
            'date': r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
            'time': r'\d{1,2}:\d{2}(:\d{2})?(\s?(AM|PM))?',
            'phone': r'^\+?1?[-.\s]?\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}$',
            'email': r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
            'integer': r'^-?\d+$',
            'decimal': r'^-?\d+\.\d+$'
        }
        
        # Patrones para detectar modelos de datos
        self.model_patterns = {
            'financial': ['revenue', 'cost', 'profit', 'expense', 'budget', 'income', 'tax'],
            'inventory': ['stock', 'quantity', 'item', 'product', 'warehouse', 'supplier'],
            'schedule': ['date', 'time', 'event', 'meeting', 'deadline', 'appointment'],
            'sales': ['customer', 'order', 'price', 'discount', 'commission', 'territory'],
            'hr': ['employee', 'salary', 'department', 'position', 'hire', 'performance']
        }
    
    async def analyze_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Análisis completo de archivo Excel
        """
        try:
            start_time = datetime.now()
            
            # Análisis con múltiples librerías
            workbook_analysis = await self._analyze_workbook_structure(file_path)
            data_analysis = await self._analyze_data_patterns(file_path)
            formula_analysis = await self._analyze_formulas(file_path)
            model_analysis = await self._detect_data_models(file_path)
            quality_analysis = await self._assess_spreadsheet_quality(file_path)
            
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return {
                "document_type": "excel",
                "analysis_results": {
                    "workbook": workbook_analysis,
                    "data_patterns": data_analysis,
                    "formulas": formula_analysis,
                    "data_models": model_analysis,
                    "quality_assessment": quality_analysis
                },
                "processing_time": processing_time,
                "status": "completed",
                "capabilities_used": self._get_analysis_capabilities()
            }
            
        except Exception as e:
            self.logger.error(f"Excel file analysis failed: {e}")
            return {
                "document_type": "excel",
                "status": "failed",
                "error": str(e),
                "processing_time": (datetime.now() - start_time).total_seconds() if 'start_time' in locals() else 0
            }
    
    async def _analyze_workbook_structure(self, file_path: str) -> Dict[str, Any]:
        """
        Analizar estructura del libro de Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            worksheets_info = []
            total_cells_with_data = 0
            
            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                
                # Determinar dimensiones reales
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Calcular rango de datos reales
                data_cells = 0
                for row in sheet.iter_rows(max_row=max_row, max_col=max_col):
                    for cell in row:
                        if cell.value is not None:
                            data_cells += 1
                
                total_cells_with_data += data_cells
                
                # Detectar si tiene headers
                has_headers = await self._detect_headers(sheet)
                
                # Analizar tipos de datos
                data_types = await self._analyze_sheet_data_types(sheet)
                
                # Estadísticas básicas
                statistics = await self._calculate_sheet_statistics(sheet)
                
                worksheet_info = {
                    "name": sheet_name,
                    "index": idx,
                    "dimensions": (max_row, max_col),
                    "data_cells": data_cells,
                    "data_density": data_cells / (max_row * max_col) if max_row * max_col > 0 else 0,
                    "has_headers": has_headers,
                    "data_types": data_types,
                    "statistics": statistics,
                    "is_active": sheet == workbook.active
                }
                
                worksheets_info.append(worksheet_info)
            
            workbook.close()
            
            return {
                "worksheets": worksheets_info,
                "total_worksheets": len(worksheets_info),
                "total_data_cells": total_cells_with_data,
                "workbook_complexity": self._calculate_workbook_complexity(worksheets_info)
            }
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _detect_headers(self, sheet) -> bool:
        """
        Detectar si la primera fila contiene encabezados
        """
        try:
            if sheet.max_row < 2:
                return False
            
            first_row_values = []
            second_row_values = []
            
            for col in range(1, min(sheet.max_column + 1, 10)):  # Primeras 10 columnas
                first_cell = sheet.cell(row=1, column=col).value
                second_cell = sheet.cell(row=2, column=col).value
                
                if first_cell is not None:
                    first_row_values.append(str(first_cell))
                if second_cell is not None:
                    second_row_values.append(str(second_cell))
            
            if not first_row_values or not second_row_values:
                return False
            
            # Heurísticas para detectar headers
            # 1. Primera fila es texto, segunda fila es números
            first_row_numeric = sum(1 for val in first_row_values if self._is_numeric(val))
            second_row_numeric = sum(1 for val in second_row_values if self._is_numeric(val))
            
            if first_row_numeric == 0 and second_row_numeric > 0:
                return True
            
            # 2. Primera fila contiene palabras típicas de headers
            header_keywords = ['name', 'id', 'date', 'amount', 'quantity', 'price', 'total', 'description']
            first_row_text = ' '.join(first_row_values).lower()
            
            keyword_matches = sum(1 for keyword in header_keywords if keyword in first_row_text)
            if keyword_matches >= 2:
                return True
            
            return False
            
        except Exception:
            return False
    
    async def _analyze_sheet_data_types(self, sheet) -> Dict[str, Any]:
        """
        Analizar tipos de datos en cada columna de la hoja
        """
        try:
            column_types = {}
            
            for col in range(1, min(sheet.max_column + 1, 50)):  # Máximo 50 columnas
                column_values = []
                
                # Recopilar valores de la columna (ignorar primera fila si es header)
                start_row = 2 if await self._detect_headers(sheet) else 1
                
                for row in range(start_row, min(sheet.max_row + 1, 1000)):  # Máximo 1000 filas
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        column_values.append(str(cell_value))
                
                if not column_values:
                    continue
                
                # Determinar tipo predominante
                type_counts = {}
                for pattern_name, pattern in self.data_patterns.items():
                    matches = sum(1 for val in column_values if re.match(pattern, val.strip()))
                    if matches > 0:
                        type_counts[pattern_name] = matches / len(column_values)
                
                # Agregar análisis de tipo nativo de openpyxl
                native_types = {}
                for row in range(start_row, min(start_row + 100, sheet.max_row + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        value_type = type(cell.value).__name__
                        native_types[value_type] = native_types.get(value_type, 0) + 1
                
                # Determinar tipo final
                if type_counts:
                    dominant_type = max(type_counts, key=type_counts.get)
                    confidence = type_counts[dominant_type]
                else:
                    dominant_type = max(native_types, key=native_types.get) if native_types else "text"
                    confidence = native_types.get(dominant_type, 0) / sum(native_types.values()) if native_types else 0
                
                column_letter = get_column_letter(col)
                column_types[column_letter] = {
                    "type": dominant_type,
                    "confidence": confidence,
                    "sample_values": column_values[:5],
                    "native_types": native_types
                }
            
            return column_types
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _calculate_sheet_statistics(self, sheet) -> Dict[str, Any]:
        """
        Calcular estadísticas de la hoja
        """
        try:
            numeric_values = []
            text_values = []
            date_values = []
            formula_count = 0
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if isinstance(cell.value, (int, float)):
                            numeric_values.append(float(cell.value))
                        elif isinstance(cell.value, str):
                            text_values.append(cell.value)
                        elif isinstance(cell.value, (date, datetime)):
                            date_values.append(cell.value)
                    
                    # Contar fórmulas
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        formula_count += 1
            
            statistics = {
                "numeric_cells": len(numeric_values),
                "text_cells": len(text_values),
                "date_cells": len(date_values),
                "formula_cells": formula_count,
                "total_data_cells": len(numeric_values) + len(text_values) + len(date_values)
            }
            
            # Estadísticas numéricas
            if numeric_values:
                statistics["numeric_stats"] = {
                    "min": min(numeric_values),
                    "max": max(numeric_values),
                    "mean": sum(numeric_values) / len(numeric_values),
                    "median": statistics.median(numeric_values) if len(numeric_values) > 1 else numeric_values[0],
                    "std_dev": statistics.stdev(numeric_values) if len(numeric_values) > 1 else 0
                }
            
            return statistics
            
        except Exception as e:
            return {"error": str(e)}
    
    def _calculate_workbook_complexity(self, worksheets_info: List[Dict]) -> str:
        """
        Calcular nivel de complejidad del libro de Excel
        """
        try:
            total_sheets = len(worksheets_info)
            total_data_cells = sum(sheet["data_cells"] for sheet in worksheets_info)
            total_formulas = sum(sheet["statistics"].get("formula_cells", 0) for sheet in worksheets_info)
            
            complexity_score = 0
            
            # Factores de complejidad
            if total_sheets > 5:
                complexity_score += 2
            elif total_sheets > 2:
                complexity_score += 1
            
            if total_data_cells > 10000:
                complexity_score += 2
            elif total_data_cells > 1000:
                complexity_score += 1
            
            if total_formulas > 100:
                complexity_score += 2
            elif total_formulas > 10:
                complexity_score += 1
            
            # Determinar nivel
            if complexity_score >= 5:
                return "high"
            elif complexity_score >= 3:
                return "medium"
            else:
                return "low"
                
        except Exception:
            return "unknown"
    
    async def _analyze_data_patterns(self, file_path: str) -> Dict[str, Any]:
        """
        Analizar patrones de datos usando pandas
        """
        try:
            # Leer con pandas para análisis avanzado
            excel_file = pd.ExcelFile(file_path)
            
            patterns_analysis = {}
            
            for sheet_name in excel_file.sheet_names[:5]:  # Máximo 5 hojas
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    sheet_patterns = {
                        "shape": df.shape,
                        "columns": list(df.columns),
                        "dtypes": df.dtypes.to_dict(),
                        "missing_values": df.isnull().sum().to_dict(),
                        "unique_values": {col: df[col].nunique() for col in df.columns},
                        "correlations": {},
                        "outliers": {},
                        "patterns": {}
                    }
                    
                    # Análisis de correlaciones (solo columnas numéricas)
                    numeric_cols = df.select_dtypes(include=[np.number]).columns
                    if len(numeric_cols) > 1:
                        corr_matrix = df[numeric_cols].corr()
                        # Encontrar correlaciones fuertes (>0.7 o <-0.7)
                        strong_correlations = []
                        for i, col1 in enumerate(numeric_cols):
                            for j, col2 in enumerate(numeric_cols[i+1:], i+1):
                                corr_value = corr_matrix.iloc[i, j]
                                if abs(corr_value) > 0.7:
                                    strong_correlations.append({
                                        "column1": col1,
                                        "column2": col2,
                                        "correlation": round(corr_value, 3)
                                    })
                        
                        sheet_patterns["correlations"] = strong_correlations
                    
                    # Detección de outliers
                    for col in numeric_cols:
                        Q1 = df[col].quantile(0.25)
                        Q3 = df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        
                        outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)][col]
                        if len(outliers) > 0:
                            sheet_patterns["outliers"][col] = {
                                "count": len(outliers),
                                "percentage": round(len(outliers) / len(df) * 100, 2),
                                "sample_values": outliers.head(5).tolist()
                            }
                    
                    # Detección de patrones específicos
                    patterns = await self._detect_specific_patterns(df)
                    sheet_patterns["patterns"] = patterns
                    
                    patterns_analysis[sheet_name] = sheet_patterns
                    
                except Exception as e:
                    patterns_analysis[sheet_name] = {"error": str(e)}
            
            return patterns_analysis
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _detect_specific_patterns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Detectar patrones específicos en los datos
        """
        patterns = {}
        
        try:
            # Patrón de serie temporal
            date_cols = df.select_dtypes(include=['datetime64']).columns
            if len(date_cols) > 0:
                patterns["time_series"] = {
                    "date_columns": list(date_cols),
                    "date_range": {
                        col: {
                            "start": str(df[col].min()),
                            "end": str(df[col].max()),
                            "span_days": (df[col].max() - df[col].min()).days
                        }
                        for col in date_cols
                    }
                }
            
            # Patrón de crecimiento/tendencia
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            for col in numeric_cols[:5]:  # Máximo 5 columnas
                if len(df[col].dropna()) > 3:
                    values = df[col].dropna().values
                    # Calcular tendencia simple
                    x = np.arange(len(values))
                    trend = np.polyfit(x, values, 1)[0]  # Pendiente
                    
                    if abs(trend) > 0.01:  # Umbral mínimo
                        patterns[f"trend_{col}"] = {
                            "slope": round(trend, 4),
                            "direction": "increasing" if trend > 0 else "decreasing",
                            "strength": "strong" if abs(trend) > 1 else "weak"
                        }
            
            # Patrón de distribución
            for col in numeric_cols[:3]:  # Máximo 3 columnas
                if len(df[col].dropna()) > 10:
                    values = df[col].dropna()
                    skewness = values.skew()
                    kurtosis = values.kurtosis()
                    
                    distribution_type = "normal"
                    if abs(skewness) > 1:
                        distribution_type = "skewed"
                    elif kurtosis > 3:
                        distribution_type = "heavy_tailed"
                    elif kurtosis < -1:
                        distribution_type = "light_tailed"
                    
                    patterns[f"distribution_{col}"] = {
                        "type": distribution_type,
                        "skewness": round(skewness, 3),
                        "kurtosis": round(kurtosis, 3)
                    }
            
            return patterns
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _analyze_formulas(self, file_path: str) -> Dict[str, Any]:
        """
        Analizar fórmulas en el archivo Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            all_formulas = []
            formula_summary = {
                "total_formulas": 0,
                "formula_types": {},
                "complexity_levels": {"simple": 0, "medium": 0, "complex": 0},
                "dependency_analysis": {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_formulas = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if hasattr(cell, 'data_type') and cell.data_type == 'f' and cell.value:
                            formula_info = await self._analyze_single_formula(cell, sheet_name)
                            sheet_formulas.append(formula_info)
                            all_formulas.append(formula_info)
                
                if sheet_formulas:
                    formula_summary["dependency_analysis"][sheet_name] = {
                        "formula_count": len(sheet_formulas),
                        "complex_formulas": sum(1 for f in sheet_formulas if f["complexity"] == "complex")
                    }
            
            # Resumir tipos y complejidad
            for formula in all_formulas:
                formula_type = formula["formula_type"]
                complexity = formula["complexity"]
                
                formula_summary["formula_types"][formula_type] = formula_summary["formula_types"].get(formula_type, 0) + 1
                formula_summary["complexity_levels"][complexity] += 1
            
            formula_summary["total_formulas"] = len(all_formulas)
            
            workbook.close()
            
            return {
                "formulas": all_formulas[:100],  # Máximo 100 fórmulas para el response
                "summary": formula_summary
            }
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _analyze_single_formula(self, cell, sheet_name: str) -> Dict[str, Any]:
        """
        Analizar una fórmula individual
        """
        try:
            formula = cell.value
            cell_address = f"{sheet_name}!{cell.coordinate}"
            
            # Clasificar tipo de fórmula
            formula_type = self._classify_formula_type(formula)
            
            # Calcular complejidad
            complexity = self._calculate_formula_complexity(formula)
            
            # Extraer dependencias (referencias a otras celdas)
            dependencies = self._extract_formula_dependencies(formula)
            
            return {
                "cell": cell_address,
                "formula": formula,
                "formula_type": formula_type,
                "complexity": complexity,
                "dependencies": dependencies,
                "is_array_formula": cell.coordinate in getattr(cell.parent, 'array_formulas', {})
            }
            
        except Exception as e:
            return {
                "cell": f"{sheet_name}!{cell.coordinate}",
                "formula": str(cell.value) if cell.value else "",
                "error": str(e)
            }
    
    def _classify_formula_type(self, formula: str) -> str:
        """
        Clasificar tipo de fórmula
        """
        formula_upper = formula.upper()
        
        if any(func in formula_upper for func in ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN']):
            return "statistical"
        elif any(func in formula_upper for func in ['IF', 'AND', 'OR', 'NOT']):
            return "logical"
        elif any(func in formula_upper for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
            return "lookup"
        elif any(func in formula_upper for func in ['DATE', 'TODAY', 'NOW', 'YEAR', 'MONTH']):
            return "datetime"
        elif any(func in formula_upper for func in ['LEFT', 'RIGHT', 'MID', 'CONCATENATE', 'LEN']):
            return "text"
        elif any(func in formula_upper for func in ['PMT', 'PV', 'FV', 'RATE', 'NPV']):
            return "financial"
        else:
            return "other"
    
    def _calculate_formula_complexity(self, formula: str) -> str:
        """
        Calcular complejidad de la fórmula
        """
        complexity_score = 0
        
        # Contar funciones
        function_count = len(re.findall(r'[A-Z]+\(', formula))
        complexity_score += function_count
        
        # Contar referencias a celdas
        cell_refs = len(re.findall(r'[A-Z]+\d+', formula))
        complexity_score += cell_refs * 0.5
        
        # Contar operadores
        operators = len(re.findall(r'[+\-*/&<>=]', formula))
        complexity_score += operators * 0.3
        
        # Detectar fórmulas anidadas
        nested_level = formula.count('(') - formula.count(')')
        complexity_score += abs(nested_level) * 2
        
        if complexity_score < 5:
            return "simple"
        elif complexity_score < 15:
            return "medium"
        else:
            return "complex"
    
    def _extract_formula_dependencies(self, formula: str) -> List[str]:
        """
        Extraer dependencias de la fórmula (referencias a otras celdas)
        """
        # Patrón para referencias de celda (ej: A1, $B$2, Sheet1!C3)
        cell_pattern = r'(?:[A-Za-z_][A-Za-z0-9_]*!)?(?:\$?[A-Z]+\$?\d+)'
        dependencies = re.findall(cell_pattern, formula)
        
        # Patrón para rangos (ej: A1:B10)
        range_pattern = r'(?:[A-Za-z_][A-Za-z0-9_]*!)?(?:\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)'
        ranges = re.findall(range_pattern, formula)
        
        return list(set(dependencies + ranges))
    
    async def _detect_data_models(self, file_path: str) -> Dict[str, Any]:
        """
        Detectar modelos de datos complejos en el Excel
        """
        try:
            models_detected = []
            
            # Leer todas las hojas para análisis
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names[:3]:  # Máximo 3 hojas
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    if df.empty:
                        continue
                    
                    # Analizar contenido para detectar modelo
                    model_type = await self._classify_data_model(df, sheet_name)
                    
                    if model_type != "unknown":
                        key_metrics = await self._extract_key_metrics(df, model_type)
                        relationships = await self._detect_data_relationships(df)
                        
                        complexity_score = self._calculate_model_complexity(df, key_metrics, relationships)
                        confidence = self._calculate_detection_confidence(df, model_type)
                        
                        model = {
                            "sheet_name": sheet_name,
                            "model_type": model_type,
                            "key_metrics": key_metrics,
                            "relationships": relationships,
                            "complexity_score": complexity_score,
                            "confidence": confidence,
                            "data_shape": df.shape
                        }
                        
                        models_detected.append(model)
                
                except Exception as e:
                    self.logger.warning(f"Failed to analyze sheet {sheet_name}: {e}")
                    continue
            
            return {
                "models_detected": models_detected,
                "total_models": len(models_detected),
                "dominant_model": max(models_detected, key=lambda x: x["confidence"])["model_type"] if models_detected else "unknown"
            }
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _classify_data_model(self, df: pd.DataFrame, sheet_name: str) -> str:
        """
        Clasificar el tipo de modelo de datos
        """
        try:
            # Analizar nombres de columnas y contenido
            column_text = ' '.join(str(col).lower() for col in df.columns)
            sheet_text = sheet_name.lower()
            
            # Combinar con una muestra del contenido
            sample_content = ""
            for col in df.columns[:5]:
                sample_values = df[col].dropna().head(10).astype(str)
                sample_content += ' '.join(sample_values.values)
            
            combined_text = (column_text + ' ' + sheet_text + ' ' + sample_content).lower()
            
            # Calcular puntuación para cada modelo
            model_scores = {}
            for model_type, keywords in self.model_patterns.items():
                score = sum(1 for keyword in keywords if keyword in combined_text)
                if score > 0:
                    model_scores[model_type] = score
            
            if model_scores:
                return max(model_scores, key=model_scores.get)
            else:
                return "unknown"
                
        except Exception:
            return "unknown"
    
    async def _extract_key_metrics(self, df: pd.DataFrame, model_type: str) -> List[str]:
        """
        Extraer métricas clave según el tipo de modelo
        """
        key_metrics = []
        
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            
            # Métricas específicas por tipo de modelo
            if model_type == "financial":
                financial_keywords = ['revenue', 'cost', 'profit', 'expense', 'budget', 'total', 'amount']
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in financial_keywords):
                        key_metrics.append(str(col))
            
            elif model_type == "sales":
                sales_keywords = ['price', 'quantity', 'order', 'customer', 'discount', 'commission']
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in sales_keywords):
                        key_metrics.append(str(col))
            
            elif model_type == "inventory":
                inventory_keywords = ['stock', 'quantity', 'item', 'product', 'warehouse']
                for col in df.columns:
                    if any(keyword in str(col).lower() for keyword in inventory_keywords):
                        key_metrics.append(str(col))
            
            # Si no hay métricas específicas, usar columnas numéricas principales
            if not key_metrics and len(numeric_cols) > 0:
                key_metrics = list(numeric_cols[:5])  # Máximo 5 columnas
            
            return key_metrics[:10]  # Máximo 10 métricas
            
        except Exception:
            return []
    
    async def _detect_data_relationships(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Detectar relaciones entre columnas de datos
        """
        relationships = []
        
        try:
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            
            if len(numeric_cols) > 1:
                # Análisis de correlación
                corr_matrix = df[numeric_cols].corr()
                
                for i, col1 in enumerate(numeric_cols):
                    for j, col2 in enumerate(numeric_cols[i+1:], i+1):
                        correlation = corr_matrix.iloc[i, j]
                        
                        if abs(correlation) > 0.5:  # Correlación significativa
                            relationship_type = "positive_correlation" if correlation > 0 else "negative_correlation"
                            strength = "strong" if abs(correlation) > 0.8 else "moderate"
                            
                            relationships.append({
                                "column1": col1,
                                "column2": col2,
                                "relationship_type": relationship_type,
                                "strength": strength,
                                "correlation_value": round(correlation, 3)
                            })
            
            # Detectar relaciones jerárquicas (suma/total)
            for col in numeric_cols:
                col_sum = df[col].sum()
                other_cols = [c for c in numeric_cols if c != col]
                
                for other_col in other_cols:
                    other_sum = df[other_col].sum()
                    if abs(col_sum - other_sum) / max(col_sum, other_sum, 1) < 0.05:  # 5% tolerance
                        relationships.append({
                            "column1": col,
                            "column2": other_col,
                            "relationship_type": "potential_sum_relationship",
                            "strength": "high",
                            "difference_percentage": abs(col_sum - other_sum) / max(col_sum, other_sum, 1) * 100
                        })
            
            return relationships[:20]  # Máximo 20 relaciones
            
        except Exception:
            return []
    
    def _calculate_model_complexity(self, df: pd.DataFrame, key_metrics: List[str], relationships: List[Dict]) -> float:
        """
        Calcular puntuación de complejidad del modelo (0-1)
        """
        try:
            complexity_score = 0.0
            
            # Factor: tamaño de datos
            data_size_factor = min(df.shape[0] * df.shape[1] / 10000, 0.3)
            complexity_score += data_size_factor
            
            # Factor: número de métricas clave
            metrics_factor = min(len(key_metrics) / 20, 0.2)
            complexity_score += metrics_factor
            
            # Factor: número de relaciones
            relationships_factor = min(len(relationships) / 10, 0.2)
            complexity_score += relationships_factor
            
            # Factor: diversidad de tipos de datos
            unique_dtypes = len(df.dtypes.unique())
            dtype_factor = min(unique_dtypes / 10, 0.15)
            complexity_score += dtype_factor
            
            # Factor: fórmulas (si se pudieran detectar)
            # Este factor se añadiría con análisis de fórmulas
            
            return min(complexity_score, 1.0)
            
        except Exception:
            return 0.5
    
    def _calculate_detection_confidence(self, df: pd.DataFrame, model_type: str) -> float:
        """
        Calcular confianza en la detección del modelo (0-1)
        """
        try:
            if model_type == "unknown":
                return 0.0
            
            confidence = 0.5  # Base confidence
            
            # Factor: coincidencia de keywords
            keywords = self.model_patterns.get(model_type, [])
            column_text = ' '.join(str(col).lower() for col in df.columns)
            
            keyword_matches = sum(1 for keyword in keywords if keyword in column_text)
            keyword_factor = min(keyword_matches / len(keywords), 0.3)
            confidence += keyword_factor
            
            # Factor: estructura de datos apropiada
            if model_type in ["financial", "sales"] and len(df.select_dtypes(include=[np.number]).columns) > 2:
                confidence += 0.2
            
            return min(confidence, 1.0)
            
        except Exception:
            return 0.0
    
    async def _assess_spreadsheet_quality(self, file_path: str) -> Dict[str, Any]:
        """
        Evaluar calidad general de la hoja de cálculo
        """
        try:
            quality_analysis = {
                "data_quality": {},
                "structure_quality": {},
                "formula_quality": {},
                "overall_score": 0.0,
                "recommendations": []
            }
            
            # Analizar calidad de datos
            data_quality = await self._assess_data_quality(file_path)
            quality_analysis["data_quality"] = data_quality
            
            # Analizar calidad de estructura
            structure_quality = await self._assess_structure_quality(file_path)
            quality_analysis["structure_quality"] = structure_quality
            
            # Analizar calidad de fórmulas
            formula_quality = await self._assess_formula_quality(file_path)
            quality_analysis["formula_quality"] = formula_quality
            
            # Calcular puntuación general
            scores = [
                data_quality.get("score", 0.5),
                structure_quality.get("score", 0.5),
                formula_quality.get("score", 0.5)
            ]
            overall_score = sum(scores) / len(scores)
            quality_analysis["overall_score"] = round(overall_score, 3)
            
            # Generar recomendaciones
            recommendations = self._generate_quality_recommendations(quality_analysis)
            quality_analysis["recommendations"] = recommendations
            
            return quality_analysis
            
        except Exception as e:
            return {"error": str(e)}
    
    async def _assess_data_quality(self, file_path: str) -> Dict[str, Any]:
        """
        Evaluar calidad de los datos
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            quality_issues = []
            quality_score = 1.0
            
            for sheet_name in excel_file.sheet_names[:3]:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                if df.empty:
                    continue
                
                # Verificar valores faltantes
                missing_percentage = (df.isnull().sum().sum() / (df.shape[0] * df.shape[1])) * 100
                if missing_percentage > 20:
                    quality_issues.append(f"High missing data in {sheet_name}: {missing_percentage:.1f}%")
                    quality_score -= 0.2
                
                # Verificar duplicados
                duplicates = df.duplicated().sum()
                if duplicates > 0:
                    quality_issues.append(f"Found {duplicates} duplicate rows in {sheet_name}")
                    quality_score -= 0.1
                
                # Verificar inconsistencia en tipos de datos
                for col in df.columns:
                    unique_types = df[col].dropna().apply(type).nunique()
                    if unique_types > 2:  # Más de 2 tipos diferentes
                        quality_issues.append(f"Inconsistent data types in column '{col}' of {sheet_name}")
                        quality_score -= 0.05
            
            return {
                "score": max(0.0, quality_score),
                "issues": quality_issues,
                "missing_data_threshold": 20,
                "duplicate_tolerance": 0
            }
            
        except Exception as e:
            return {"error": str(e), "score": 0.5}
    
    async def _assess_structure_quality(self, file_path: str) -> Dict[str, Any]:
        """
        Evaluar calidad de la estructura
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            structure_issues = []
            structure_score = 1.0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Verificar si hay headers
                has_headers = await self._detect_headers(sheet)
                if not has_headers and sheet.max_row > 1:
                    structure_issues.append(f"No clear headers detected in {sheet_name}")
                    structure_score -= 0.2
                
                # Verificar estructura de datos
                data_density = 0
                total_cells = sheet.max_row * sheet.max_column
                filled_cells = 0
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            filled_cells += 1
                
                data_density = filled_cells / total_cells if total_cells > 0 else 0
                
                if data_density < 0.1:  # Menos del 10% de celdas con datos
                    structure_issues.append(f"Low data density in {sheet_name}: {data_density:.1%}")
                    structure_score -= 0.1
            
            workbook.close()
            
            return {
                "score": max(0.0, structure_score),
                "issues": structure_issues,
                "min_data_density": 0.1
            }
            
        except Exception as e:
            return {"error": str(e), "score": 0.5}
    
    async def _assess_formula_quality(self, file_path: str) -> Dict[str, Any]:
        """
        Evaluar calidad de las fórmulas
        """
        try:
            formula_analysis = await self._analyze_formulas(file_path)
            formula_issues = []
            formula_score = 1.0
            
            if "summary" in formula_analysis:
                summary = formula_analysis["summary"]
                
                # Verificar complejidad excesiva
                complex_formulas = summary["complexity_levels"].get("complex", 0)
                total_formulas = summary["total_formulas"]
                
                if total_formulas > 0:
                    complex_ratio = complex_formulas / total_formulas
                    if complex_ratio > 0.3:  # Más del 30% son complejas
                        formula_issues.append(f"High ratio of complex formulas: {complex_ratio:.1%}")
                        formula_score -= 0.2
                
                # Verificar diversidad de tipos de fórmula
                formula_types = len(summary["formula_types"])
                if formula_types == 1 and total_formulas > 10:
                    formula_issues.append("Limited variety in formula types")
                    formula_score -= 0.1
            
            return {
                "score": max(0.0, formula_score),
                "issues": formula_issues,
                "complex_formula_threshold": 0.3
            }
            
        except Exception as e:
            return {"error": str(e), "score": 0.5}
    
    def _generate_quality_recommendations(self, quality_analysis: Dict) -> List[str]:
        """
        Generar recomendaciones de calidad
        """
        recommendations = []
        
        # Recomendaciones de datos
        data_quality = quality_analysis.get("data_quality", {})
        if data_quality.get("score", 1.0) < 0.7:
            recommendations.append("Address data quality issues: missing values and duplicates")
        
        # Recomendaciones de estructura
        structure_quality = quality_analysis.get("structure_quality", {})
        if structure_quality.get("score", 1.0) < 0.7:
            recommendations.append("Improve spreadsheet structure: add clear headers and organize data")
        
        # Recomendaciones de fórmulas
        formula_quality = quality_analysis.get("formula_quality", {})
        if formula_quality.get("score", 1.0) < 0.7:
            recommendations.append("Simplify complex formulas and add documentation")
        
        # Recomendación general
        overall_score = quality_analysis.get("overall_score", 0.5)
        if overall_score > 0.8:
            recommendations.append("Spreadsheet quality is excellent - maintain current standards")
        elif overall_score < 0.5:
            recommendations.append("Consider restructuring the spreadsheet for better maintainability")
        
        return recommendations if recommendations else ["No specific recommendations - quality assessment incomplete"]
    
    def _is_numeric(self, value: str) -> bool:
        """Verificar si un valor es numérico"""
        try:
            float(str(value).replace(',', '').replace('$', '').strip())
            return True
        except (ValueError, AttributeError):
            return False
    
    def _get_analysis_capabilities(self) -> Dict[str, bool]:
        """Obtener capacidades de análisis disponibles"""
        return {
            "openpyxl": True,
            "pandas": True,
            "xlrd": True,
            "data_pattern_analysis": True,
            "formula_analysis": True,
            "data_model_detection": True,
            "quality_assessment": True,
            "correlation_analysis": True,
            "outlier_detection": True,
            "statistical_analysis": True
        }
    
    async def extract_data_for_conversion(self, file_path: str) -> Dict[str, Any]:
        """
        Extraer datos estructurados para conversión a otros formatos
        """
        try:
            analysis = await self.analyze_excel_file(file_path)
            
            # Extraer datos estructurados
            structured_data = {
                "worksheets": [],
                "metadata": {},
                "data_models": [],
                "formulas": []
            }
            
            # Leer datos de cada hoja
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                worksheet_data = {
                    "name": sheet_name,
                    "data": df.to_dict(orient='records'),
                    "columns": list(df.columns),
                    "shape": df.shape,
                    "dtypes": df.dtypes.to_dict()
                }
                
                structured_data["worksheets"].append(worksheet_data)
            
            # Agregar análisis
            structured_data["metadata"] = analysis.get("analysis_results", {}).get("workbook", {})
            structured_data["data_models"] = analysis.get("analysis_results", {}).get("data_models", {})
            structured_data["formulas"] = analysis.get("analysis_results", {}).get("formulas", {})
            
            return {
                "status": "success",
                "structured_data": structured_data,
                "analysis_summary": analysis
            }
            
        except Exception as e:
            return {
                "status": "failed",
                "error": str(e)
            }